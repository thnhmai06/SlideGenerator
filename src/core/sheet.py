"""Module for handling spreadsheet workbooks and sheets."""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from src.config import SPREADSHEET_EXTENSIONS
from src.core.exceptions import FileExtensionNotSupported, IndexOutOfRange


class SheetGroup:
    """Represents a workbook file."""

    def __init__(self, file_path: Path):
        ext = Path(file_path).suffix.lstrip('.')
        if ext not in SPREADSHEET_EXTENSIONS:
            raise FileExtensionNotSupported(ext)

        self._workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        self._tables = {sheet_name: SheetTable(self._workbook, sheet_name) for sheet_name in self._workbook.sheetnames}

    @property
    def workbook(self):
        """Get the workbook object."""
        return self._workbook

    @property
    def tables(self):
        """Get all tables in the workbook."""
        return self._tables

    @property
    def file_path(self):
        """Get the file path of the workbook."""
        return self._workbook.path

    @property
    def name(self):
        """Get the name of the workbook file."""
        return self._workbook.properties.title


class SheetTable:
    """Represents a table in a sheet (in workbook file)."""

    def __init__(self, workbook: Workbook, sheet_name: str):
        self._sheet = workbook[sheet_name]

        for row in self._sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if self._min_row is None or cell.row < self._min_row:
                        self._min_row = cell.row
                    if self._max_row is None or cell.row > self._max_row:
                        self._max_row = cell.row
                    if self._min_col is None or cell.column < self._min_col:
                        self._min_col = cell.column
                    if self._max_col is None or cell.column > self._max_col:
                        self._max_col = cell.column

        self._headers = [
            self._sheet.cell(row=self._min_row, column=col).value
            for col in range(self._min_col, self._max_col + 1)
        ]

    @property
    def sheet(self):
        """Get the sheet object."""
        return self._sheet

    @property
    def headers(self) -> list[str]:
        """Get the headers of the table."""
        return self._headers

    @property
    def amounts(self) -> int:
        """Get the number of data rows in the table."""
        return self._max_row - self._min_row

    @property
    def name(self) -> str:
        """Get the name of the sheet."""
        return self._sheet.title

    def get_row(self, index: int):
        """Get a row by index (1-based)."""
        if index < 1 or index > self.amounts:
            raise IndexOutOfRange(index, (1, self.amounts))

        index += self._min_row  # Adjust index to actual row number

        row_data: dict[str, str] = {}
        for col in range(self._min_col, self._max_col + 1):
            header = self.headers[col - self._min_col]
            cell_value = self._sheet.cell(row=index, column=col).value
            row_data[header] = cell_value
        return row_data
